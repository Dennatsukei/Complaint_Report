"""Read Excel reports and reviews into unified complaint records."""

import re
import pandas as pd
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from fields import AGGREGATED_COLUMNS
from transforms import extract_incident_date


class ComplaintExtractor:
    def __init__(self, file_path):
        self.file_path = Path(file_path)

        self.wb = load_workbook(
            self.file_path,
            data_only=True
        )

        self.ws = self.wb.active

    # Daily
    def extract_daily(self):
        """
        Extract a daily report.

        Daily report:
        - report date = incident date
        - complaints are located below the complaint section
        """
        report_start_date, report_end_date = self.extract_report_date()

        complaints = self._extract_complaint_section(
            incident_date=report_start_date
        )

        complaints = self._add_report_period(
            complaints,
            report_start_date,
            report_end_date
        )

        return {
            "report_start_date": report_start_date,
            "report_end_date": report_end_date,
            "report_type": "daily",
            "source_file": self.file_path.name,
            "complaints": complaints
        }

    # Weekly
    def extract_weekly(self):
        """
        Extract a weekly report.

        Weekly report:
        - complaints are stored in a table
        - incident date comes from the date column
        """
        report_start_date, report_end_date = self.extract_report_date()

        complaints = self._extract_event_table()

        complaints = self._add_report_period(
            complaints,
            report_start_date,
            report_end_date
        )

        return {
            "report_start_date": report_start_date,
            "report_end_date": report_end_date,
            "report_type": "weekly",
            "source_file": self.file_path.name,
            "complaints": complaints
        }

    # Monthly
    def extract_monthly(self):
        """
        Extract a monthly report.

        Monthly report:
        - report date is a period
        - incident date is extracted from complaint content
        """
        report_start_date, report_end_date = self.extract_report_date()

        complaints = self._extract_complaint_section(
            report_start_date=report_start_date,
            report_end_date=report_end_date
        )

        complaints = self._add_report_period(
            complaints,
            report_start_date,
            report_end_date
        )

        return {
            "report_start_date": report_start_date,
            "report_end_date": report_end_date,
            "report_type": "monthly",
            "source_file": self.file_path.name,
            "complaints": complaints
        }

    # Platform reviews
    def extract_reviews(self):
        """
        Extract platform reviews.

        Expected columns:
        - 日期
        - 评价内容
        - 平台
        - 评分
        - 房型
        """

        df = pd.read_excel(self.file_path)

        records = []

        for index, row in df.iterrows():

            records.append({
                "incident_date": row["日期"],
                "raw_content": row["评价内容"],
                "platform": row["平台"],
                "score": row["评分"],
                "room_type": row["房型"],
                "review_id": index + 1,
                "source_path": self.file_path.name,
            })

        return records

    # Report date
    def extract_report_date(self):
        """
        Extract report date / date period.

        Supported:
        日期：2026.08.19
        日期：2026.08.19-2026.08.20
        日期：2026.08.19-08.20
        日期：2026.08.19~2026.08.20
        日期：2026.08.19至2026.08.20
        """
        pattern = (
            r"日期[：:]\s*"
            r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})"
            r"(?:\s*[-~至到]\s*"
            r"(?:(\d{4})[./-])?"
            r"(\d{1,2})[./-](\d{1,2}))?"
        )

        for row in self.ws.iter_rows():

            for cell in row:

                if cell.value is None:
                    continue

                if isinstance(cell.value, datetime):

                    report_date = cell.value.date()

                    return report_date, report_date

                text = str(cell.value).strip()

                match = re.search(
                    pattern,
                    text
                )

                if not match:
                    continue

                (
                    year1,
                    month1,
                    day1,
                    year2,
                    month2,
                    day2
                ) = match.groups()

                start_date = date(
                    int(year1),
                    int(month1),
                    int(day1)
                )

                if month2 is None:
                    return start_date, start_date

                if year2 is None:
                    year2 = year1

                end_date = date(
                    int(year2),
                    int(month2),
                    int(day2)
                )

                return start_date, end_date

        return None, None

    # Daily / Monthly complaint section
    def _find_complaint_section(self):
        """
        Find the complaint section.

        Supported section titles:
        - 客诉及处理方案
        - 投诉事件及处理方案
        """
        keywords = [
            "客诉及处理方案",
            "投诉事件及处理方案"
        ]

        for row in self.ws.iter_rows():

            for cell in row:

                if cell.value is None:
                    continue

                text = str(cell.value).strip()

                for keyword in keywords:

                    if keyword in text:
                        return cell.row

        return None

    def _extract_complaint_section(
        self,
        incident_date=None,
        report_start_date=None,
        report_end_date=None
    ):
        """
        Extract complaints from a numbered complaint section.

        One numbered cell = one complaint event.
        The content is NOT split here.

        If report dates are provided, incident date is extracted
        from the complaint content.
        """
        start_row = self._find_complaint_section()

        if start_row is None:
            return []

        complaints = []

        for row in self.ws.iter_rows(
            min_row=start_row + 1
        ):

            for cell in row:

                if cell.value is None:
                    continue

                text = str(cell.value).strip()

                match = re.match(
                    r"^(\d+)\s*[、.。]\s*(.*)",
                    text,
                    re.S
                )

                if not match:
                    continue

                complaint_id = int(
                    match.group(1)
                )

                content = match.group(2).strip()

                if (
                    report_start_date is not None
                    or report_end_date is not None
                ):
                    extracted_incident_date = extract_incident_date(
                        content,
                        report_start_date,
                        report_end_date
                    )
                else:
                    extracted_incident_date = incident_date

                complaints.append({
                    "complaint_id": complaint_id,
                    "incident_date": extracted_incident_date,
                    "raw_content": content
                })

        return complaints

    # Weekly event table
    def _find_event_table_header(self):
        """
        Find the header row of the weekly event table.
        """
        for row in self.ws.iter_rows():

            date_col = None
            content_col = None

            for cell in row:

                if cell.value is None:
                    continue

                text = str(cell.value).strip()

                if text == "日期":
                    date_col = cell.column

                elif text == "投诉及发生事件":
                    content_col = cell.column

            if date_col and content_col:

                return {
                    "row": row[0].row,
                    "date_col": date_col,
                    "content_col": content_col
                }

        return None

    def _extract_event_table(self):
        """
        Extract complaints from the weekly event table.
        """
        header = self._find_event_table_header()

        if header is None:
            return []

        header_row = header["row"]
        date_col = header["date_col"]
        content_col = header["content_col"]

        complaints = []

        complaint_id = 1

        for row in self.ws.iter_rows(
            min_row=header_row + 1
        ):

            row_number = row[0].row

            date_cell = self.ws.cell(
                row=row_number,
                column=date_col
            )

            content_cell = self.ws.cell(
                row=row_number,
                column=content_col
            )

            if content_cell.value is None:
                continue

            content = str(
                content_cell.value
            ).strip()

            if not content:
                continue

            incident_date = self._parse_event_date(
                date_cell.value
            )

            complaints.append({
                "complaint_id": complaint_id,
                "incident_date": incident_date,
                "raw_content": content
            })

            complaint_id += 1

        return complaints

    # Weekly incident date
    def _parse_event_date(self, value):
        """
        Parse a date from the weekly event table.

        Supported:
        8.1
        8.11
        08.01
        2026.8.1
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        text = str(value).strip()

        if not text:
            return None

        # Full date
        match = re.fullmatch(
            r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})",
            text
        )

        if match:

            year, month, day = map(
                int,
                match.groups()
            )

            return date(
                year,
                month,
                day
            )

        # Month + day
        match = re.fullmatch(
            r"(\d{1,2})[./-](\d{1,2})",
            text
        )

        if match:

            month, day = map(
                int,
                match.groups()
            )

            report_start_date, _ = (
                self.extract_report_date()
            )

            if report_start_date is None:
                return None

            return date(
                report_start_date.year,
                month,
                day
            )

        return None

    # Metadata
    def _add_report_period(
        self,
        complaints,
        report_start_date,
        report_end_date
    ):
        """
        Attach report period to every complaint.
        """
        for complaint in complaints:

            complaint["report_start_date"] = (
                report_start_date
            )

            complaint["report_end_date"] = (
                report_end_date
            )

        return complaints


class ComplaintAggregator:

    def __init__(self):
        self.records = []

    def add_report(self, result):
        """
        Add a daily, weekly, or monthly report.
        """

        report_type = result["report_type"]
        source_file = result["source_file"]

        for complaint in result["complaints"]:

            self.records.append({
                "incident_date": complaint["incident_date"],
                "report_start_date": complaint[
                    "report_start_date"
                ],
                "report_end_date": complaint[
                    "report_end_date"
                ],
                "raw_content": complaint["raw_content"],
                "source": "internal_report",
                "source_file": source_file,
                "report_type": report_type,
                "complaint_id": complaint["complaint_id"],

                # Not applicable to internal reports
                "platform": None,
                "score": None,
                "room_type": None,
            })

    def add_platform_reviews(self, reviews):
        """
        Add platform review records.
        """

        for review in reviews:

            self.records.append({
                "incident_date": review["incident_date"],
                "report_start_date": review["incident_date"],
                "report_end_date": review["incident_date"],
                "raw_content": review["raw_content"],
                "source": "platform_review",
                "source_file": review["source_path"],
                "report_type": "platform_review",
                "complaint_id": review["review_id"],

                "platform": review["platform"],
                "score": review["score"],
                "room_type": review["room_type"],
            })

    def to_dataframe(self):
        """
        Convert all records into a DataFrame
        with a stable column order.
        """

        return pd.DataFrame(
            self.records,
            columns=AGGREGATED_COLUMNS
        )

    def save(self, output_path):
        """
        Save aggregated records.
        """

        df = self.to_dataframe()

        df.to_csv(
            output_path,
            index=False,
            encoding="utf-8-sig"
        )

        return df
